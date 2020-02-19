import openpyxl
from openpyxl import Workbook
import random

#產生隨機資料
def generateRandomInformation(filename):
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.append(['姓名','課程','成績'])
    #中文名字中的第一、第二、第三個字
    first = tuple('趙錢孫李')
    middle = tuple('偉昀琛東')
    last = tuple('坤豔志')
    #課程名稱
    subjects = ('語文','數學','英語')
    #隨機產生200筆資料
    for i in range(200):
        line = []
        r = random.randint(1,100)
        name = random.choice(first)
        #按一定機率產生只有兩個字的中文名字
        if r>50:
            name = name + random.choice(middle)
        name = name + random.choice(last)
        #依序產生姓名、課程名稱和成績
        line.append(name)
        line.append(random.choice(subjects))
        line.append(random.randint(0,100))
        worksheet.append(line)
    #保存資料，產生Excel 2007格式的檔案
    workbook.save(filename)

def getResult(oldfile, newfile):
    #用來存放結果資料的字典
    result = dict()
    #開啟原始資料
    workbook = openpyxl.load_workbook(oldfile)
    worksheet = workbook.worksheets[0]
    #遍歷原始資料
    #跳過第0列的表頭
    for row in worksheet.rows[1:]:
        #姓名，課程名稱，本次成績
        name, subject, grade= row[0].value, row[1].value, row[2].value
        #取得目前姓名對應的課程名稱和成績
        #如果result字典沒有此姓名，則返回空字典
        t = result.get(name, {})
        #取得學生目前課程的成績，若不存在，返回0
        f = t.get(subject, 0)
        #只保留學生該課程的最高成績
        if grade > f:
            t[subject] = grade
            result[name] = t
    #建立Excel檔案
    workbook1 = Workbook()
    worksheet1 = workbook1.worksheets[0]
    worksheet1.append(['姓名','課程','成績'])
    #將result字典的結果寫入Excel檔案
    for name, t in result.items():
        for subject, grade in t.items():
            worksheet1.append([name, subject, grade])
    workbook1.save(newfile)

if __name__ == '__main__':
    oldfile = r'D:\test.xlsx'
    newfile = r'D:\result.xlsx'
    generateRandomInformation(oldfile)
    getResult(oldfile, newfile)

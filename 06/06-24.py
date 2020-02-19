import openpyxl
from openpyxl import Workbook
fn = r'f:\test.xlsx'				#檔案名稱
wb = Workbook()						#建立工作簿
ws = wb.create_sheet(title='你好，世界')	#建立工作表
ws['A1'] = '這是第一個儲存格' 		#設定儲存格
ws['B1'] = 3.1415926
wb.save(fn)							#儲存Excel檔案
wb = openpyxl.load_workbook(fn)		#開啟既有的Excel檔案
ws = wb.worksheets[1]				#開啟指定索引的工作表
print(ws['A1'].value) 				#讀取並輸出指定儲存格的值
ws.append([1,2,3,4,5])				#加入一列資料
ws.merge_cells('F2:F3')				#合併儲存格
ws['F2'] = "=sum(A2:E2)" 			#寫入公式
for r in range(10,15):
    for c in range(3,8):
        _ = ws.cell(row=r, column=c, value=r*c)	#寫入儲存格資料
wb.save(fn)
